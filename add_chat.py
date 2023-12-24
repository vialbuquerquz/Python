from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

#lendo pasta de trabalho
wb = load_workbook("data/pivotable.xlsx")

sheet = wb["Relatório"]

#Indicando mínimo e máximo de colunas da planilha
min_column = wb.active.min_column
max_column = wb.active.max_column

#INdicando mínimo e máximo de linhas da planilha
min_row = wb.active.min_row
max_row = wb.active.max_row

#Adicionando dados e categorias no gráfico
barchart = BarChart()

data = Reference(
    sheet,
    min_col = min_column + 1,
    max_col = max_column,
    min_row = min_row,
    max_row = max_row
)

categories = Reference(
    sheet,
    min_col=min_column,
    max_col=min_column,
    min_row=min_row + 1,
    max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#Criando o gráfico
sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por fabricantes"
barchart.style = 2

#Salvando workbook
wb.save("data/barchart.xlsx")